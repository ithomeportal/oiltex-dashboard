#!/usr/bin/env python3
"""
Import historical Argus data from new Excel pack (March 2026).

Files:
  - Argus Historical Prices(3).xlsx — WTL Midland + WTI Midland (NEW data)
  - Argus Historical Prices.xlsx / (4).xlsx — NYMEX Futures + CMA (fill gaps)
  - Argus Historical Prices(1).xlsx / (2).xlsx — WTI Houston (fill gaps)

This script:
1. Updates argus_pricing rows with missing WTL Midland + Midland diff data
2. Fills missing oil_prices entries (NYMEX settle, WTI Houston)
3. Recalculates est_net_daily and est_net_mtd
"""

import os
import sys
import openpyxl
import psycopg2
from datetime import datetime

# --- DB connection ---
DB_CONFIG = {
    'host': os.environ.get('DATABASE_HOST', 'pg-111cab4b-unlkdata.b.aivencloud.com'),
    'port': int(os.environ.get('DATABASE_PORT', '10261')),
    'user': os.environ.get('DATABASE_USER', ''),
    'password': os.environ.get('DATABASE_PASSWORD', ''),
    'dbname': os.environ.get('DATABASE_NAME', 'defaultdb'),
    'sslmode': 'require',
}

PICTURES_DIR = os.path.expanduser('~/Pictures')

def parse_contract_month(timing: str) -> str:
    """Convert 'Apr-2026' to '2026-04'"""
    if not timing:
        return None
    months = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
              'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
    parts = str(timing).split('-')
    if len(parts) != 2:
        return None
    mon, year = parts[0][:3], parts[1]
    return f"{year}-{months.get(mon, '01')}"


def parse_date(dt) -> str:
    """Convert datetime to 'YYYY-MM-DD'"""
    if isinstance(dt, datetime):
        return dt.strftime('%Y-%m-%d')
    return str(dt)[:10]


def safe_float(val):
    """Convert to float or None"""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def load_file3_wtl_midland():
    """Load WTL Midland + WTI Midland data from File (3)"""
    path = os.path.join(PICTURES_DIR, 'Argus Historical Prices(3).xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb['Price History']

    rows = []
    for r in range(3, ws.max_row + 1):
        date = ws.cell(r, 1).value
        timing = ws.cell(r, 2).value
        if not date or not timing:
            continue

        rows.append({
            'report_date': parse_date(date),
            'contract_month': parse_contract_month(timing),
            # WTL Midland month 1 flat prices
            'wtl_midland_low': safe_float(ws.cell(r, 6).value),     # Price Low
            'wtl_midland_high': safe_float(ws.cell(r, 8).value),    # Price High
            # WTL Midland weighted average month 1
            'wtl_midland_wtd_avg': safe_float(ws.cell(r, 14).value),  # Price
            'midland_diff_daily': safe_float(ws.cell(r, 13).value),   # Diff (WTL Midland wtd avg diff)
            # WTL Midland MTD weighted average month 1
            'midland_diff_mtd': safe_float(ws.cell(r, 19).value),     # Diff
        })

    wb.close()
    print(f"Loaded {len(rows)} rows from File (3) WTL Midland")
    return rows


def load_nymex_data():
    """Load NYMEX Futures settlement from File (0) or (4)"""
    path = os.path.join(PICTURES_DIR, 'Argus Historical Prices.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb['Price History']

    rows = []
    for r in range(3, ws.max_row + 1):
        date = ws.cell(r, 1).value
        settlement = safe_float(ws.cell(r, 6).value)  # NYMEX WTI futures month 1 Settlement
        if not date or settlement is None:
            continue
        rows.append({
            'report_date': parse_date(date),
            'nymex_settlement': settlement,
        })

    wb.close()
    print(f"Loaded {len(rows)} rows from NYMEX futures data")
    return rows


def load_wti_houston_data():
    """Load WTI Houston weighted average from File (1)"""
    path = os.path.join(PICTURES_DIR, 'Argus Historical Prices(1).xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb['Price History']

    rows = []
    for r in range(3, ws.max_row + 1):
        date = ws.cell(r, 1).value
        diff = safe_float(ws.cell(r, 3).value)    # WTI Houston wtd avg month 1 Diff
        price = safe_float(ws.cell(r, 4).value)   # WTI Houston wtd avg month 1 Price
        if not date:
            continue
        rows.append({
            'report_date': parse_date(date),
            'wti_houston_diff': diff,
            'wti_houston_price': price,
        })

    wb.close()
    print(f"Loaded {len(rows)} rows from WTI Houston data")
    return rows


def import_data(dry_run: bool = False):
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # --- Phase 1: Update argus_pricing with WTL Midland + Midland diff ---
    print("\n=== Phase 1: WTL Midland + Midland diff → argus_pricing ===")
    wtl_rows = load_file3_wtl_midland()

    updated = 0
    inserted = 0
    for row in wtl_rows:
        # First check if row exists
        cur.execute(
            "SELECT id, nymex_cma_td, cma_diff_daily, cma_diff_mtd, midland_diff_daily FROM argus_pricing WHERE report_date = %s AND contract_month = %s",
            (row['report_date'], row['contract_month'])
        )
        existing = cur.fetchone()

        if existing:
            ex_id, cma_td, cma_d, cma_mtd, mid_d = existing

            # Recalculate est_net with the new midland data
            est_net_daily = None
            est_net_mtd = None
            if cma_td is not None and cma_d is not None and row['midland_diff_daily'] is not None:
                est_net_daily = round(float(cma_td) + float(cma_d) + row['midland_diff_daily'], 4)
            if cma_td is not None and cma_mtd is not None and row['midland_diff_mtd'] is not None:
                est_net_mtd = round(float(cma_td) + float(cma_mtd) + row['midland_diff_mtd'], 4)

            cur.execute("""
                UPDATE argus_pricing SET
                    wtl_midland_low = COALESCE(%s, wtl_midland_low),
                    wtl_midland_high = COALESCE(%s, wtl_midland_high),
                    wtl_midland_wtd_avg = COALESCE(%s, wtl_midland_wtd_avg),
                    midland_diff_daily = COALESCE(%s, midland_diff_daily),
                    midland_diff_mtd = COALESCE(%s, midland_diff_mtd),
                    est_net_daily = COALESCE(%s, est_net_daily),
                    est_net_mtd = COALESCE(%s, est_net_mtd)
                WHERE id = %s
            """, (
                row['wtl_midland_low'], row['wtl_midland_high'], row['wtl_midland_wtd_avg'],
                row['midland_diff_daily'], row['midland_diff_mtd'],
                est_net_daily, est_net_mtd,
                ex_id
            ))
            updated += 1
            status = "UPDATED" if mid_d is None else "KEPT (already had midland)"
            print(f"  {row['report_date']} ({row['contract_month']}): {status} wtl_avg={row['wtl_midland_wtd_avg']} mid_d={row['midland_diff_daily']} mid_mtd={row['midland_diff_mtd']} est_net_d={est_net_daily} est_net_mtd={est_net_mtd}")
        else:
            # Row doesn't exist — shouldn't happen since CMA was already imported
            print(f"  {row['report_date']} ({row['contract_month']}): MISSING in DB (no CMA row)")

    print(f"Phase 1 complete: {updated} rows updated")

    # --- Phase 2: Fill missing NYMEX settlement in oil_prices ---
    print("\n=== Phase 2: NYMEX Futures Settlement → oil_prices ===")
    nymex_rows = load_nymex_data()

    nymex_inserted = 0
    for row in nymex_rows:
        cur.execute(
            "SELECT id FROM oil_prices WHERE date = %s AND source = 'NYMEX' AND price_type = 'WTI_FUTURES_SETTLE'",
            (row['report_date'],)
        )
        if not cur.fetchone():
            cur.execute("""
                INSERT INTO oil_prices (date, source, price_type, value, unit)
                VALUES (%s, 'NYMEX', 'WTI_FUTURES_SETTLE', %s, 'USD/bbl')
                ON CONFLICT (date, source, price_type) DO NOTHING
            """, (row['report_date'], row['nymex_settlement']))
            nymex_inserted += 1
            print(f"  {row['report_date']}: INSERTED NYMEX settle = {row['nymex_settlement']}")
        else:
            pass  # already exists

    print(f"Phase 2 complete: {nymex_inserted} new NYMEX entries")

    # --- Phase 3: Fill missing WTI Houston in oil_prices ---
    print("\n=== Phase 3: WTI Houston → oil_prices ===")
    houston_rows = load_wti_houston_data()

    houston_inserted = 0
    for row in houston_rows:
        for price_type, value in [
            ('WTI_HOUSTON_WTD_AVG', row['wti_houston_price']),
            ('WTI_HOUSTON_DIFF', row['wti_houston_diff']),
        ]:
            if value is None:
                continue
            cur.execute(
                "SELECT id FROM oil_prices WHERE date = %s AND source = 'ARGUS' AND price_type = %s",
                (row['report_date'], price_type)
            )
            if not cur.fetchone():
                cur.execute("""
                    INSERT INTO oil_prices (date, source, price_type, value, unit)
                    VALUES (%s, 'ARGUS', %s, %s, 'USD/bbl')
                    ON CONFLICT (date, source, price_type) DO NOTHING
                """, (row['report_date'], price_type, value))
                houston_inserted += 1
                print(f"  {row['report_date']}: INSERTED {price_type} = {value}")

    print(f"Phase 3 complete: {houston_inserted} new WTI Houston entries")

    # --- Commit or rollback ---
    if dry_run:
        print("\n*** DRY RUN — rolling back ***")
        conn.rollback()
    else:
        conn.commit()
        print("\n*** All changes committed ***")

    cur.close()
    conn.close()


if __name__ == '__main__':
    dry_run = '--dry-run' in sys.argv
    if dry_run:
        print("=== DRY RUN MODE ===\n")
    import_data(dry_run=dry_run)
