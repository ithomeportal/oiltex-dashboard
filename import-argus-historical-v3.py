#!/usr/bin/env python3
"""
Import historical Argus data from "Argus Historicals for WTI WTL CMA Diff.xlsx".

This file contains 3 data groups with ~247 rows (2025-03-10 to 2026-03-06):
  1. WTI Diff to CMA Nymex month 1 (cols 0-5): Date, Timing, Diff Low, Diff High, Diff, Diff. Change
  2. WTI Midland month 1 (cols 6-14): Timing, Diff Low, Diff High, Diff, Price Low, Price High, Price, Change, Diff. Change
  3. WTL Midland month 1 (cols 15-23): Timing, Diff Low, Diff High, Diff, Price Low, Price High, Price, Change, Diff. Change

Maps to:
  - argus_pricing: nymex_cma_td, cma_diff_daily, midland_diff_daily (WTL Midland diff),
                   wtl_midland_low/high/wtd_avg, est_net_daily
  - oil_prices: ARGUS/WTI_MIDLAND_DIFF (from WTI Midland section)
"""

import os
import sys
import openpyxl
import psycopg2
from datetime import datetime

DB_CONFIG = {
    'host': os.environ.get('DATABASE_HOST', 'pg-111cab4b-unlkdata.b.aivencloud.com'),
    'port': int(os.environ.get('DATABASE_PORT', '10261')),
    'user': os.environ.get('DATABASE_USER', ''),
    'password': os.environ.get('DATABASE_PASSWORD', ''),
    'dbname': os.environ.get('DATABASE_NAME', 'defaultdb'),
    'sslmode': 'require',
}

EXCEL_PATH = os.path.expanduser('~/Pictures/Argus Historicals for WTI WTL CMA Diff.xlsx')

MONTHS = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
           'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}


def parse_contract_month(timing):
    if not timing:
        return None
    parts = str(timing).split('-')
    if len(parts) != 2:
        return None
    mon, year = parts[0][:3], parts[1]
    return f"{year}-{MONTHS.get(mon, '01')}"


def parse_date(dt):
    if isinstance(dt, datetime):
        return dt.strftime('%Y-%m-%d')
    return str(dt)[:10]


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def load_data():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Price History']

    rows = []
    for r in range(3, ws.max_row + 1):
        date = ws.cell(r, 1).value
        timing = ws.cell(r, 2).value  # Contract month (same across all 3 groups)
        if not date or not timing:
            continue

        # WTI Diff to CMA Nymex (cols 1-6, 1-indexed)
        cma_diff_daily = safe_float(ws.cell(r, 5).value)    # Diff (weighted avg)

        # WTI Midland (cols 7-15, 1-indexed)
        wti_midland_diff = safe_float(ws.cell(r, 10).value)   # Diff
        wti_midland_price = safe_float(ws.cell(r, 13).value)  # Price (flat)

        # WTL Midland (cols 16-24, 1-indexed)
        wtl_midland_diff = safe_float(ws.cell(r, 19).value)   # Diff
        wtl_midland_price_low = safe_float(ws.cell(r, 20).value)   # Price Low
        wtl_midland_price_high = safe_float(ws.cell(r, 21).value)  # Price High
        wtl_midland_price = safe_float(ws.cell(r, 22).value)       # Price (wtd avg)

        # Derive NYMEX CMA TD: WTI Midland Price - WTI Midland Diff - CMA Diff
        # Because: WTI Midland Price = NYMEX settle + WTI Midland Diff
        #          NYMEX settle = NYMEX CMA TD + CMA Diff
        #          So: NYMEX CMA TD = WTI Midland Price - WTI Midland Diff - CMA Diff
        nymex_cma_td = None
        if wti_midland_price is not None and wti_midland_diff is not None and cma_diff_daily is not None:
            nymex_cma_td = round(wti_midland_price - wti_midland_diff - cma_diff_daily, 4)

        # est_net_daily = nymex_cma_td + cma_diff_daily + wtl_midland_diff (= WTL Midland flat price)
        est_net_daily = None
        if nymex_cma_td is not None and cma_diff_daily is not None and wtl_midland_diff is not None:
            est_net_daily = round(nymex_cma_td + cma_diff_daily + wtl_midland_diff, 4)

        rows.append({
            'report_date': parse_date(date),
            'contract_month': parse_contract_month(timing),
            'nymex_cma_td': nymex_cma_td,
            'cma_diff_daily': cma_diff_daily,
            'midland_diff_daily': wtl_midland_diff,  # DB field = WTL Midland diff
            'wtl_midland_low': wtl_midland_price_low,
            'wtl_midland_high': wtl_midland_price_high,
            'wtl_midland_wtd_avg': wtl_midland_price,
            'est_net_daily': est_net_daily,
            'wti_midland_diff': wti_midland_diff,  # For oil_prices table
        })

    wb.close()
    print(f"Loaded {len(rows)} rows from Excel")
    return rows


def import_data(dry_run=False):
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    rows = load_data()

    # --- Phase 1: Upsert argus_pricing ---
    print("\n=== Phase 1: argus_pricing (nymex_cma_td, cma_diff, midland_diff, wtl prices, est_net) ===")
    ap_inserted = 0
    ap_updated = 0

    for row in rows:
        result = cur.execute("""
            INSERT INTO argus_pricing (
                report_date, contract_month, nymex_cma_td,
                cma_diff_daily, midland_diff_daily,
                est_net_daily,
                wtl_midland_low, wtl_midland_high, wtl_midland_wtd_avg,
                extraction_confidence, raw_extraction
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 100, %s)
            ON CONFLICT (report_date, contract_month)
            DO UPDATE SET
                nymex_cma_td = COALESCE(EXCLUDED.nymex_cma_td, argus_pricing.nymex_cma_td),
                cma_diff_daily = COALESCE(EXCLUDED.cma_diff_daily, argus_pricing.cma_diff_daily),
                midland_diff_daily = COALESCE(EXCLUDED.midland_diff_daily, argus_pricing.midland_diff_daily),
                est_net_daily = COALESCE(EXCLUDED.est_net_daily, argus_pricing.est_net_daily),
                wtl_midland_low = COALESCE(EXCLUDED.wtl_midland_low, argus_pricing.wtl_midland_low),
                wtl_midland_high = COALESCE(EXCLUDED.wtl_midland_high, argus_pricing.wtl_midland_high),
                wtl_midland_wtd_avg = COALESCE(EXCLUDED.wtl_midland_wtd_avg, argus_pricing.wtl_midland_wtd_avg)
        """, (
            row['report_date'], row['contract_month'], row['nymex_cma_td'],
            row['cma_diff_daily'], row['midland_diff_daily'],
            row['est_net_daily'],
            row['wtl_midland_low'], row['wtl_midland_high'], row['wtl_midland_wtd_avg'],
            '{"source": "argus_historical_excel_v3", "file": "Argus Historicals for WTI WTL CMA Diff.xlsx"}'
        ))

        # Check if insert or update (use rowcount approach)
        cur.execute(
            "SELECT id FROM argus_pricing WHERE report_date = %s AND contract_month = %s",
            (row['report_date'], row['contract_month'])
        )
        existing = cur.fetchone()
        if existing:
            ap_updated += 1
        else:
            ap_inserted += 1

        print(f"  {row['report_date']} ({row['contract_month']}): cma_td={row['nymex_cma_td']} cma_d={row['cma_diff_daily']} mid_d={row['midland_diff_daily']} wtl={row['wtl_midland_wtd_avg']} est_net={row['est_net_daily']}")

    print(f"Phase 1 complete: {ap_inserted} inserted, {ap_updated} updated")

    # --- Phase 2: Upsert oil_prices with WTI_MIDLAND_DIFF ---
    print("\n=== Phase 2: oil_prices (ARGUS/WTI_MIDLAND_DIFF) ===")
    op_inserted = 0

    for row in rows:
        if row['wti_midland_diff'] is None:
            continue
        cur.execute(
            "SELECT id FROM oil_prices WHERE date = %s AND source = 'ARGUS' AND price_type = 'WTI_MIDLAND_DIFF'",
            (row['report_date'],)
        )
        if not cur.fetchone():
            cur.execute("""
                INSERT INTO oil_prices (date, source, price_type, value, unit)
                VALUES (%s, 'ARGUS', 'WTI_MIDLAND_DIFF', %s, 'USD/bbl')
                ON CONFLICT (date, source, price_type) DO NOTHING
            """, (row['report_date'], row['wti_midland_diff']))
            op_inserted += 1
            print(f"  {row['report_date']}: INSERTED WTI_MIDLAND_DIFF = {row['wti_midland_diff']}")

    print(f"Phase 2 complete: {op_inserted} new WTI_MIDLAND_DIFF entries")

    if dry_run:
        print("\n*** DRY RUN — rolling back ***")
        conn.rollback()
    else:
        conn.commit()
        print("\n*** All changes committed ***")

    cur.close()
    conn.close()


if __name__ == '__main__':
    dry_run = '--dry-run' in sys.argv
    if dry_run:
        print("=== DRY RUN MODE ===\n")
    import_data(dry_run=dry_run)
